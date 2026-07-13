# -*- coding: utf-8 -*-
import openpyxl, io, json, re
XLSX = r"C:/Scripts/AiEmployeeTest/AiEmployeeTest.xlsx"
OUTJS = r"C:/Scripts/AiEmployeeTest/js/questions.js"
OUTTSV = r"C:/Scripts/AiEmployeeTest/backend/answers_with_tags.tsv"

# tag map from docx (verified option order matches xlsx for profiled blocks)
TAGS = {
 101:{'A':'opponent','B':'neutral','C':'enthusiast'},
 102:{'A':'opponent','B':'neutral','C':'enthusiast'},
 103:{'A':'risk-security','B':'risk-quality','C':'neutral'},
 104:{'A':'opponent','B':'pragmatic','C':'enthusiast'},
 105:{'A':'personal-responsibility','B':'shared-responsibility','C':'neutral'},
 106:{'A':'anxious','B':'confident','C':'enthusiast'},
 107:{'A':'cautious','B':'neutral','C':'enthusiast'},
 108:{'A':'opponent','B':'pragmatic','C':'neutral'},
 109:{'A':'anxious','B':'neutral','C':'enthusiast'},
 110:{'A':'pragmatic','B':'pragmatic','C':'skeptic'},
 201:{'A':'high-interest','B':'low-interest','C':'situational'},
 202:{'A':'high-interest','B':'situational','C':'low-interest'},
 203:{'A':'practical-writing','B':'practical-research','C':'low-interest'},
 204:{'A':'workshop','B':'self-paced','C':'flexible'},
 205:{'A':'high-interest','B':'situational','C':'low-interest'},
 206:{'A':'integration-critical','B':'integration-nice','C':'integration-indifferent'},
 207:{'A':'high-interest','B':'situational','C':'low-interest'},
 208:{'A':'high-interest','B':'situational','C':'low-interest'},
 209:{'A':'deep-interest','B':'practical-only','C':'low-interest'},
 210:{'A':'high-interest','B':'situational','C':'low-interest'},
 411:{'A':'risk-high','B':'risk-aware','C':'risk-avoidant'},
 511:{'A':'level-0','B':'level-1','C':'level-2','D':'level-3'},
}

def block_of(qid):
    p=str(qid)[0]
    return {'1':'attitude','2':'interest','3':'knowledge','4':'security','5':'self'}[p]
MASCOT={'attitude':'neutral','interest':'happy','knowledge':'thinking','security':'thinking','self':'neutral'}

wb=openpyxl.load_workbook(XLSX, data_only=True)

# --- question text from "Questions" sheet (pairs of columns) ---
qtext={}
ws=wb['Questions']
rows=list(ws.iter_rows(values_only=True))
for r in rows[1:]:
    for ci in (0,2,4):
        if ci+1 < len(r) and r[ci] not in (None,'') and r[ci+1] not in (None,''):
            try: qid=int(float(r[ci]))
            except: continue
            qtext[qid]=str(r[ci+1]).strip()

# --- options + correct from "Answers" sheet ---
ans={}
wa=wb['Answers']
arows=list(wa.iter_rows(values_only=True))
for r in arows[1:]:
    if r[0] in (None,''): continue
    try: qid=int(float(r[0]))
    except: continue
    opts={'A':r[1],'B':r[2],'C':r[3],'D':r[4]}
    correct=(str(r[5]).strip() if len(r)>5 and r[5] not in (None,'') else '')
    ans[qid]={'opts':{k:(str(v).strip() if v not in (None,'') else '') for k,v in opts.items()},'correct':correct}

order=[511]+list(range(101,111))+list(range(201,211))+list(range(301,311))+[411]
def is_own(txt): return 'вой вариант' in (txt or '')

questions=[]
for qid in order:
    blk=block_of(qid)
    a=ans.get(qid,{'opts':{},'correct':''})
    opts=a['opts']; correct=a['correct']
    q={'id':str(qid),'block':blk,'mascot':MASCOT[blk],'text':qtext.get(qid,'')}
    if blk=='knowledge':
        q['type']='knowledge'
        q['options']=[{'key':k,'text':opts[k]} for k in ('A','B','C','D') if opts.get(k)]
        q['correct']=correct
    elif blk=='self':
        q['type']='self'
        q['options']=[{'key':k,'text':opts[k],'tag':TAGS[qid].get(k,'')} for k in ('A','B','C','D') if opts.get(k)]
    else: # profile (attitude/interest/security)
        q['type']='profile'
        real=[k for k in ('A','B','C','D') if opts.get(k) and not is_own(opts[k])]
        q['options']=[{'key':k,'text':opts[k],'tag':TAGS.get(qid,{}).get(k,'')} for k in real]
        q['allowOwn']=any(is_own(opts.get(k,'')) for k in ('A','B','C','D'))
    questions.append(q)

# sanity
assert len(questions)==32, len(questions)
kn=[q for q in questions if q['type']=='knowledge']
assert len(kn)==10 and all(q['correct'] in ('A','B','C','D') for q in kn), [q['correct'] for q in kn]

header=("/* AUTO-GENERATED from AiEmployeeTest.xlsx + docx tags. Do not edit by hand;\n"
        "   regenerate via backend/gen.py. In production questions come from the sheet. */\n")
js=header+"window.DEMO_QUESTIONS = "+json.dumps(questions, ensure_ascii=False, indent=2)+";\n"
with io.open(OUTJS,'w',encoding='utf-8') as f: f.write(js)

# --- TSV block: paste into Answers sheet as tag columns (G:J -> Тег A..D) ---
with io.open(OUTTSV,'w',encoding='utf-8') as f:
    f.write("Тег A\tТег B\tТег C\tТег D\n")
    # rows must line up with the Answers sheet data rows (101..511 order as in sheet)
    sheet_ids=[int(float(r[0])) for r in arows[1:] if r[0] not in (None,'')]
    for qid in sheet_ids:
        t=TAGS.get(qid,{})
        f.write("\t".join(t.get(k,'') for k in ('A','B','C','D'))+"\n")

# --- also write tags into local xlsx Answers sheet (columns after existing) ---
wb2=openpyxl.load_workbook(XLSX)  # keep formatting/formulas
wa2=wb2['Answers']
base=wa2.max_column
labels=['Тег A','Тег B','Тег C','Тег D']
for i,l in enumerate(labels): wa2.cell(row=1,column=base+1+i,value=l)
rr=2
for r in arows[1:]:
    if r[0] in (None,''): rr+=1; continue
    qid=int(float(r[0])); t=TAGS.get(qid,{})
    for i,k in enumerate(('A','B','C','D')):
        wa2.cell(row=rr,column=base+1+i,value=t.get(k,''))
    rr+=1
wb2.save(XLSX)

print("questions:",len(questions))
print("knowledge correct:",[ (q['id'],q['correct']) for q in kn])
print("profiled with own-answer:",sum(1 for q in questions if q.get('allowOwn')))
print("wrote:",OUTJS,"and",OUTTSV,"and updated xlsx Answers (+Тег A..D)")
